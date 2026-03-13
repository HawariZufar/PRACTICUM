# Libraries Import
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import numpy as np

# Global Variabel Declaration
STANDART_RETIREMENT_AGE = 60
RETURN_OF_INVESTMENT = 0.06
STATUS_LIST = ['Eligible', 'Standart', 'Not Eligible']

ACCOUNT_DATA = {
    'admin': 'admin123',
    'user': 'user123'
}

CLIENT_DATA_LIST = []

# Login
def login():
    max_attempts = 3
    attempts = 0
    is_logged_in = False

    while attempts < max_attempts and is_logged_in == False:
        username = input("Enter username: ").strip()
        password = input("Enter password: ").strip()

        if username not in ACCOUNT_DATA or ACCOUNT_DATA[username] != password:
            attempts += 1  
            print(f"Invalid username or password. {max_attempts - attempts} attempts remaining.")
        else:
            print('Login successful')
            is_logged_in = True
            return True

    return False

# Validate Input
def validate_input(prompt, input_type, valid_values=None, min_value=None, max_value=None):
    while True:
        try:
            if input_type == 'str':
                value = input(prompt).strip().capitalize()
                if not value:
                    raise ValueError("Input cannot be empty")
                if valid_values and value not in valid_values:
                    raise ValueError(f"Input must be one of {valid_values}")
                return value

            elif input_type == 'int':
                value = int(input(prompt))
                if min_value is not None and value < min_value:
                    raise ValueError("Value too small")
                if max_value is not None and value > max_value:
                    raise ValueError("Value too large")
                return value

            elif input_type == 'float':
                value = float(input(prompt))
                if min_value is not None and value < min_value:
                    raise ValueError("Value too small")
                return value

        except ValueError as e:
            print(f"Error: {e}")

# Input Method
def input_method():
    return validate_input(
        "Choose input method (manual/excel): ",
        'str',
        valid_values=['Manual', 'Excel']
    )

# Manual Input
def manual_input():
    n = validate_input("Enter the number of client data to input: ", 'int', min_value=1)

    for i in range(n):
        print(f"Client-{i+1}")
        CLIENT_DATA_LIST.append({
            'name': validate_input("Enter client name: ", 'str'),
            'age': validate_input("Enter client age: ", 'int', min_value=1, max_value=59),
            'salary': validate_input("Enter Monthly Salary (Rp): ", 'float', min_value=1),
            'saving_rate': validate_input("Enter savings rate: ", 'float', min_value=0.1),
            'gender': validate_input("Enter client gender (male/female): ", 'str', valid_values=['Male', 'Female']),
            'pmt_spending': validate_input("Enter PMT Spending (Rp): ", 'float', min_value=1)
        })

# Excel Input
def excel_input():
    try:
        path = input("Excel file path: ").strip().replace("\"", "")
        df = pd.read_excel(path)

        df.columns = df.columns.str.lower()
        head = ['name', 'age', 'gender', 'salary', 'saving_rate',  'pmt_spending']

        for column in head:
            if column not in df.columns:
                raise ValueError(f"Missing column: {column}")

        for i, row in df.iterrows():
            CLIENT_DATA_LIST.append({
                'name': row['name'],
                'age': int(row['age']),
                'salary': float(row['salary']),
                'saving_rate': float(row['saving_rate']),
                'gender': row['gender'].capitalize(),
                'pmt_spending': float(row['pmt_spending'])
            })

        print("Excel data imported successfully!")

    except Exception as e:
        print(f"Excel Error: {e}")

# Calculation
def calculation(client):
    years_to_retirement = STANDART_RETIREMENT_AGE - client['age']
    months_to_retirement = years_to_retirement * 12

    if client['gender'] == 'Male':
        life_expectancy = 70
    else:
        life_expectancy = 75
        
    post_retirement_time = (life_expectancy - STANDART_RETIREMENT_AGE) * 12

    r = RETURN_OF_INVESTMENT / 12

    pmt_saving = client["salary"] * (client["saving_rate"] / 100)
    
    pva = client['pmt_spending'] * ( (1 - (1 / (1 + r) ** post_retirement_time)) / r )

    fva_standard = pmt_saving * (((1 + r) ** months_to_retirement - 1) / r)

    coverage_ratio = fva_standard / pva

    fva_target = 0
    month_target = 1
    while fva_target < pva:
        fva_target = pmt_saving * ( ((1 + r) ** month_target - 1) / r )
        month_target += 1

    year_target = np.ceil(month_target / 12)
    retirement_age = (client['age']) + year_target

    pmt_saving_standard = pva * r / ((1 + r) ** months_to_retirement - 1)
    
    saving_rate_standard = (pmt_saving_standard / client['salary']) * 100

    client.update({
        'years_to_retirement': years_to_retirement,
        'life_expectancy': life_expectancy,
        'pmt_saving': pmt_saving,
        'pva': pva,
        'fva_standard': fva_standard,
        'coverage_ratio': coverage_ratio,
        'year_target': year_target,
        'retirement_age': retirement_age,
        'pmt_saving_standard': pmt_saving_standard,
        'saving_rate_standard': saving_rate_standard
    })

    return client

# Classification
def classification(client):
    if client['coverage_ratio'] > 1:
        client['status'] = STATUS_LIST[0]
    elif client['coverage_ratio'] == 1:
        client['status'] = STATUS_LIST[1]
    else:
        client['status'] = STATUS_LIST[2]
    return client

# Show Graphic
def show_graph(df):
    x = range(len(df))
    y = []

    for i in x:
        y.append(i+0.35)

    plt.figure(figsize=(10, 5))
    plt.bar(x, df['fva_standard'], 0.35, label='FVA')
    plt.bar(y, df['pva'], 0.35, label='PVA')

    plt.xticks([i + 0.35 / 2 for i in x], df['name'])
    plt.ylabel("Amount (Rp)")
    plt.title("FVA vs PVA Comparison")
    plt.legend()

    def format_ticks(value, _):
        return f"{int(value):,}"
    
    plt.gca().yaxis.set_major_formatter(FuncFormatter(format_ticks))
    plt.tight_layout()
    plt.show()

# Show Data
def show_data():  
    print("\n")
    print("=" * 126)
    print(f"| {'SUMMARY':^122} |")
    print("-" * 126)
    print(f"| {'No':<3} | {'Name':<10} | {'Age':<3} | {'Gender':<7} | {'Salary':<15} | {'PMT Spending':<15} | {'PVA':<16} | {'FVA Standard':<16} | {'Status':<13} |")
    print("=" * 126)

    for idx, client in enumerate(CLIENT_DATA_LIST, start=1):   
        print(f"| {idx:<3} | {client['name']:<10} | {client['age']:<3} | {client['gender']:<7} | Rp{client['salary']:<13,.2f} | Rp{client['pmt_spending']:<13,.2f} | Rp{client['pva']:<14,.2f} | Rp{client['fva_standard']:<14,.2f} | {client["status"]:<13} |")

    print("=" * 126)

    print("\n")
    print("=" * 123)
    print(f"| {'RECOMMENDATION':^119} |")
    print("-" * 123)

    print(f"| {'No':<3} | {'Name':<10} | {'Existing PMT':<16} | {'Existing Rate':<7} | {'Minimum PMT':<16} | {'Minimum Rate':<8} | {'Remaining Time':<12} | {'Estimated Time':<12} |")
    print("=" * 123)

    for idx, client in enumerate(CLIENT_DATA_LIST, start=1):
        print(f"| {idx:<3} | {client['name']:<10} | Rp{client['pmt_saving']:<14,.2f} | {client['saving_rate']:<12,.2f}% | Rp{client['pmt_saving_standard']:<14,.2f} | {client['saving_rate_standard']:<11,.2f}% | {client['years_to_retirement']:<8} years | {client['year_target']:<8} years |")

    print("=" * 123)

# Save To Excel
def save_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rencana Pensiun"

    header_font = Font(bold=True)
    align_center = Alignment(horizontal='center')

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.alignment = align_center

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("rencana_pensiun.xlsx")
    print("Data saved to rencana_pensiun.xlsx")


# Main
def main():
    method = input_method()

    if method == 'Manual':
        manual_input()
    else:
        excel_input()

    for i, client in enumerate(CLIENT_DATA_LIST):
        CLIENT_DATA_LIST[i] = classification(calculation(client))

    df = pd.DataFrame(CLIENT_DATA_LIST)
    show_data()
    save_excel(df)
    show_graph(df)

# Program Overview
if login():
    main()
else:
    print("Maximum login attempts reached. Program terminated.")